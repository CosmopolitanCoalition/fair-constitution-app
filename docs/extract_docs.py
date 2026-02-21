#!/usr/bin/env python3
"""
Extract text from reference documents in docs/ for Claude Code to read.
Run this at the start of any Claude Code session that needs constitutional context.

Usage: python3 docs/extract_docs.py
Output: docs/extracted/ folder with markdown versions of all reference docs.
"""

import os
import sys

docs_dir = os.path.join(os.path.dirname(__file__))
output_dir = os.path.join(docs_dir, 'extracted')
os.makedirs(output_dir, exist_ok=True)

def extract_docx(filename, output_name):
    try:
        from docx import Document
        path = os.path.join(docs_dir, filename)
        if not os.path.exists(path):
            print(f"  SKIP: {filename} not found")
            return
        doc = Document(path)
        text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
        out_path = os.path.join(output_dir, output_name)
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"  OK: {filename} → extracted/{output_name} ({len(text)} chars)")
    except ImportError:
        print("  ERROR: python-docx not installed. Run: pip install python-docx --break-system-packages")
    except Exception as e:
        print(f"  ERROR extracting {filename}: {e}")

def extract_xlsx(filename, output_name):
    try:
        import openpyxl
        path = os.path.join(docs_dir, filename)
        if not os.path.exists(path):
            print(f"  SKIP: {filename} not found")
            return
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n## Sheet: {sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                row_text = ' | '.join([str(c) if c is not None else '' for c in row])
                if row_text.strip(' |'):
                    lines.append(row_text)
        text = '\n'.join(lines)
        out_path = os.path.join(output_dir, output_name)
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"  OK: {filename} → extracted/{output_name} ({len(text)} chars)")
    except ImportError:
        print("  ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    except Exception as e:
        print(f"  ERROR extracting {filename}: {e}")

def copy_drawio(filename, output_name):
    path = os.path.join(docs_dir, filename)
    if not os.path.exists(path):
        print(f"  SKIP: {filename} not found")
        return
    import shutil
    out_path = os.path.join(output_dir, output_name)
    shutil.copy2(path, out_path)
    size = os.path.getsize(path)
    print(f"  OK: {filename} → extracted/{output_name} ({size} bytes, read as XML)")

print("Extracting reference documents...")
extract_docx('Fair_Constitution_Labeled.docx', 'fair_constitution.md')
extract_docx('CGA_Architecture_Plan.docx', 'architecture_plan.md')
extract_xlsx('CGA_Constitutional_Roles_Forms_Chart.xlsx', 'roles_forms_chart.md')
extract_xlsx('Topic_Knowledge.xlsx', 'topic_knowledge.md')
copy_drawio('The_Chart.drawio', 'the_chart.xml')
copy_drawio('App_Flows.drawio', 'app_flows.xml')
print("\nDone. Read docs/extracted/ for constitutional context.")
